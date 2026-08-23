"""Quick one-table Excel downloads (#5) — shared by the KPI-gap and
submission-status export views."""
from io import BytesIO

import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def rows_to_xlsx_response(title, headers, rows, filename):
    """Build a single-sheet xlsx HttpResponse.

    rows values may be ints (formatted #,##0), floats (rendered as-is),
    strings or None (rendered as an em dash).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    header_fill = PatternFill('solid', fgColor='1F3A5F')
    header_font = Font(color='FFFFFF', bold=True)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical='center')
    ws.freeze_panes = 'A2'

    for r, row in enumerate(rows, start=2):
        for c, value in enumerate(row, start=1):
            if value is None:
                cell = ws.cell(row=r, column=c, value='—')
            elif isinstance(value, int):
                cell = ws.cell(row=r, column=c, value=value)
                cell.number_format = '#,##0'
            else:
                cell = ws.cell(row=r, column=c, value=value)

    # Autosize with a sane cap so long headlines don't blow out the sheet
    for col in range(1, len(headers) + 1):
        width = max(
            (len(str(ws.cell(row=row_, column=col).value or '')) + 2
             for row_ in range(1, min(ws.max_row, 200) + 1)),
            default=len(headers[col - 1]) + 2,
        )
        ws.column_dimensions[get_column_letter(col)].width = max(min(width, 55), 10)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
