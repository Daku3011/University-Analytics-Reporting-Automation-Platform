import os
import uuid
import pathlib
from io import BytesIO
from unittest import skipUnless
from unittest.mock import patch, MagicMock

from django.test import TestCase, Client
from django.urls import reverse
from django.contrib.auth.models import User
from django.contrib.messages import get_messages
from django.core.files.uploadedfile import SimpleUploadedFile

from colleges.models import College, Department
from accounts.models import Profile
from analytics_app.models import MonthlyAnalytics, KpiTarget
from reports.models import MonthlyReport, QuarterlyReport, UploadedDocumentReport
from reports.views import _save_uploaded_file


def _weasyprint_available():
    from importlib.util import find_spec
    return find_spec('weasyprint') is not None

class ReportsSecurityAndRobustnessTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(username='testuser', password='password')
        self.profile = self.user.profile
        self.profile.role = 'super_admin'
        self.profile.college = None
        self.profile.save()
        self.college = College.objects.create(name='Sarvajanik College of Engineering and Technology', code='SCET')
        
    def login_user(self):
        self.client.login(username='testuser', password='password')

    def test_generate_monthly_get_redirects(self):
        self.login_user()
        response = self.client.get(reverse('generate_monthly'))
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('report_dashboard'))

    def test_generate_monthly_missing_params(self):
        self.login_user()
        response = self.client.post(reverse('generate_monthly'), {
            'college': self.college.id,
            # 'month' and 'year' are missing
        })
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('report_dashboard'))
        
        messages = list(get_messages(response.wsgi_request))
        self.assertTrue(any("Missing required parameters" in str(m) for m in messages))

    def test_generate_monthly_invalid_types(self):
        self.login_user()
        response = self.client.post(reverse('generate_monthly'), {
            'college': self.college.id,
            'month': 'not-a-number',
            'year': '2026'
        })
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('report_dashboard'))
        
        messages = list(get_messages(response.wsgi_request))
        self.assertTrue(any("must be valid numeric values" in str(m) for m in messages))

    def test_generate_monthly_invalid_bounds(self):
        self.login_user()
        response = self.client.post(reverse('generate_monthly'), {
            'college': self.college.id,
            'month': '13',
            'year': '2026'
        })
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('report_dashboard'))
        
        messages = list(get_messages(response.wsgi_request))
        self.assertTrue(any("Month must be between 1 and 12" in str(m) for m in messages))

    def test_generate_monthly_invalid_college(self):
        self.login_user()
        response = self.client.post(reverse('generate_monthly'), {
            'college': 9999,
            'month': '6',
            'year': '2026'
        })
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('report_dashboard'))
        
        messages = list(get_messages(response.wsgi_request))
        self.assertTrue(any("Specified College does not exist" in str(m) for m in messages))

    @patch('weasyprint.HTML')
    def test_generate_monthly_weasyprint_compilation_failure(self, mock_html):
        self.login_user()
        # Mock HTML().write_pdf to raise an exception
        mock_html.return_value.write_pdf.side_effect = Exception("System library GTK not found")
        
        response = self.client.post(reverse('generate_monthly'), {
            'college': self.college.id,
            'month': '6',
            'year': '2026'
        })
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('report_dashboard'))
        
        messages = list(get_messages(response.wsgi_request))
        self.assertTrue(any("PDF compilation failed" in str(m) for m in messages))

    @patch('weasyprint.HTML')
    def test_generate_monthly_success(self, mock_html):
        self.login_user()
        # Ensure we have mock analytics
        MonthlyAnalytics.objects.create(
            college=self.college, month=6, year=2026,
            instagram_views=100, facebook_views=200, total_views=300
        )
        
        response = self.client.post(reverse('generate_monthly'), {
            'college': self.college.id,
            'month': '6',
            'year': '2026'
        })
        
        # Should create a MonthlyReport and redirect to its preview
        report = MonthlyReport.objects.first()
        self.assertIsNotNone(report)
        self.assertEqual(report.college, self.college)
        self.assertEqual(report.month, 6)
        self.assertEqual(report.year, 2026)
        
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('preview_monthly', kwargs={'report_id': report.id}))

    def test_generate_quarterly_missing_params(self):
        self.login_user()
        response = self.client.post(reverse('generate_quarterly'), {
            # 'quarter' is missing
            'year': '2026'
        })
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('report_dashboard'))
        
        messages = list(get_messages(response.wsgi_request))
        self.assertTrue(any("Missing quarter parameter" in str(m) for m in messages))

    def test_generate_quarterly_invalid_bounds(self):
        self.login_user()
        response = self.client.post(reverse('generate_quarterly'), {
            'quarter': '5',
            'year': '2026'
        })
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('report_dashboard'))
        
        messages = list(get_messages(response.wsgi_request))
        self.assertTrue(any("Quarter must be between 1 and 4" in str(m) for m in messages))

    @patch('weasyprint.HTML')
    @patch('google.genai.Client')
    def test_generate_quarterly_success(self, mock_client, mock_html):
        self.login_user()
        
        # Mock Gemini response
        mock_response = MagicMock()
        mock_response.text = "<h2>Quarterly Summary HTML</h2>"
        mock_client.return_value.models.generate_content.return_value = mock_response
        
        response = self.client.post(reverse('generate_quarterly'), {
            'quarter': '1',
            'year': '2026'
        })
        
        # Should create a QuarterlyReport and redirect to its preview
        report = QuarterlyReport.objects.first()
        self.assertIsNotNone(report)
        self.assertEqual(report.quarter, 1)
        self.assertEqual(report.year, 2026)
        
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse('preview_quarterly', kwargs={'report_id': report.id}))

    def test_compare_reports_missing_params(self):
        self.login_user()
        response = self.client.post(reverse('compare_reports'), {
            'college': self.college.id,
            # month_a and month_b are missing
        })
        self.assertEqual(response.status_code, 200)
        self.assertIn("Missing required fields for comparison", response.context['comparison_html'])

    def test_compare_reports_invalid_bounds(self):
        self.login_user()
        response = self.client.post(reverse('compare_reports'), {
            'college': self.college.id,
            'month_a': '0',
            'month_b': '13',
            'year': '2026'
        })
        self.assertEqual(response.status_code, 200)
        self.assertIn("Months must be between 1 and 12", response.context['comparison_html'])

    @patch('google.genai.Client')
    def test_compare_reports_success(self, mock_client):
        self.login_user()
        
        # Mock Gemini MoM comparison response
        mock_response = MagicMock()
        mock_response.text = "<h2>Comparison Summary</h2>"
        mock_client.return_value.models.generate_content.return_value = mock_response
        
        response = self.client.post(reverse('compare_reports'), {
            'college': self.college.id,
            'month_a': '1',
            'month_b': '2',
            'year': '2026'
        })
        self.assertEqual(response.status_code, 200)
        self.assertIsNotNone(response.context['comparison_html'])
        self.assertIn("Comparison Summary", response.context['comparison_html'])

    def test_save_uploaded_file_path_traversal(self):
        # Create a mock file with a malicious name containing path traversal payloads
        malicious_filename = "../../../etc/passwd.pdf"
        file_content = b"%PDF-1.4 mock content"
        uploaded_file = SimpleUploadedFile(malicious_filename, file_content, content_type='application/pdf')
        
        # Invoke _save_uploaded_file
        save_path, rel_path = _save_uploaded_file(uploaded_file, "prefix")
        
        # The file stem must be fully sanitized and not escape the uploaded_sources boundaries
        self.assertTrue(save_path.exists())
        self.assertNotIn("..", save_path.name)
        self.assertNotIn("/", save_path.name)
        self.assertNotIn("etc", str(save_path.parent.relative_to(save_path.parent.parent)))
        self.assertTrue(save_path.name.endswith(".pdf"))
        
        # Clean up the created physical file
        if save_path.exists():
            save_path.unlink()


class PortfolioTestBase(TestCase):
    """Shared fixtures for the Annual Portfolio (#5) tests."""

    def setUp(self):
        self.client = Client()
        self.co1 = College.objects.create(name="Alpha College", code="CO1")
        self.co2 = College.objects.create(name="Beta College", code="CO2")
        MonthlyAnalytics.objects.create(
            college=self.co1, month=1, year=2026,
            instagram_views=600, facebook_views=400, total_views=1000,
            instagram_reach=700, facebook_reach=300, total_reach=1000,
            followers_gained=50, reels_count=4, graphics_count=10,
        )
        MonthlyAnalytics.objects.create(
            college=self.co1, month=2, year=2026,
            instagram_views=800, facebook_views=400, total_views=1200,
            instagram_reach=900, facebook_reach=300, total_reach=1200,
            followers_gained=70, reels_count=6, graphics_count=12,
        )
        # Previous year baseline for YoY
        MonthlyAnalytics.objects.create(
            college=self.co1, month=1, year=2025,
            instagram_views=500, facebook_views=300, total_views=800,
            instagram_reach=500, facebook_reach=300, total_reach=800,
            followers_gained=30,
        )
        # Department-scope row must never inflate institute totals
        self.dept = Department.objects.create(college=self.co1, name="Computer", code="CE")
        MonthlyAnalytics.objects.create(
            college=self.co1, department=self.dept, month=1, year=2026,
            instagram_views=99999, total_views=99999,
        )
        KpiTarget.objects.create(
            college=self.co1, year=2026, metric='total_views', target_value=2000)

        self.super_user = User.objects.create_user(username="superadmin", password="password123")
        self.super_user.profile.role = "super_admin"
        self.super_user.profile.save()
        self.college_user = User.objects.create_user(username="collegeadmin", password="password123")
        self.college_user.profile.role = "college_admin"
        self.college_user.profile.college = self.co1
        self.college_user.profile.save()


class PortfolioServiceTests(PortfolioTestBase):
    def test_context_golden_numbers(self):
        from reports.services.portfolio_service import build_portfolio_context
        ctx = build_portfolio_context(self.co1, 2026)
        ex = ctx['chapters']['executive_summary']
        views = next(m for m in ex['yoy'] if m['label'] == 'Total Views')
        self.assertEqual(views['current'], 2200)      # dept row (99999) excluded
        self.assertEqual(views['previous'], 800)
        self.assertEqual(views['change'], 175.0)
        self.assertEqual(ex['months_reported'], 2)

        kpi = ctx['chapters']['kpi_performance']['rows']
        self.assertEqual(len(kpi), 1)
        self.assertEqual(kpi[0]['actual'], 2200)
        self.assertTrue(kpi[0]['on_track'])
        self.assertEqual(ex['kpi_targets_met'], 1)
        self.assertEqual(ex['kpi_avg_attainment'], 110)


class PortfolioExportTests(PortfolioTestBase):
    def _context(self):
        from reports.services.portfolio_service import build_portfolio_context
        return build_portfolio_context(self.co1, 2026)

    @skipUnless(_weasyprint_available(), 'WeasyPrint not installed')
    def test_pdf_download(self):
        self.client.login(username="superadmin", password="password123")
        resp = self.client.get(reverse('portfolio_pdf'), {'college': self.co1.id, 'year': 2026})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/pdf')
        self.assertTrue(resp.content.startswith(b'%PDF'))

    def test_excel_workbook_reloads_with_expected_sheets(self):
        from openpyxl import load_workbook
        from reports.services.excel_service import build_portfolio_workbook

        payload = build_portfolio_workbook(self._context())
        wb = load_workbook(BytesIO(payload))
        for sheet in ['Summary', 'Social Media', 'Events', 'Media Coverage',
                      'Press Releases', 'KPI Performance']:
            self.assertIn(sheet, wb.sheetnames)
        ws = wb['KPI Performance']
        self.assertEqual(ws.cell(row=2, column=3).value, 2000)   # target
        self.assertEqual(ws.cell(row=2, column=4).value, 2200)   # actual (dept excluded)

    def test_word_document_opens_and_has_chapters(self):
        from docx import Document as ReadDocx
        from reports.services.word_service import build_portfolio_docx

        payload = build_portfolio_docx(self._context())
        doc = ReadDocx(BytesIO(payload))
        text = '\n'.join(p.text for p in doc.paragraphs)
        for chapter in ['Executive Summary', 'Social Media Performance', 'Events',
                        'Media Coverage', 'Press Releases', 'KPI Performance']:
            self.assertIn(chapter, text)


class PortfolioViewTests(PortfolioTestBase):
    def test_preview_requires_login(self):
        resp = self.client.get(reverse('portfolio_preview'))
        self.assertEqual(resp.status_code, 302)

    def test_super_admin_can_pick_college(self):
        self.client.login(username="superadmin", password="password123")
        resp = self.client.get(reverse('portfolio_preview'),
                               {'college': self.co2.id, 'year': 2026})
        self.assertContains(resp, 'Beta College')

    def test_college_admin_locked_to_own_institute(self):
        self.client.login(username="collegeadmin", password="password123")
        resp = self.client.get(reverse('portfolio_preview'),
                               {'college': self.co2.id, 'year': 2026})
        self.assertContains(resp, 'Alpha College')
        self.assertNotContains(resp, 'Beta College')

    def test_excel_and_word_downloads(self):
        self.client.login(username="superadmin", password="password123")
        url = reverse('portfolio_excel')
        resp = self.client.get(url, {'college': self.co1.id, 'year': 2026})
        self.assertIn('spreadsheetml', resp['Content-Type'])
        self.assertIn(f'CO1_Annual_Portfolio_2026.xlsx', resp['Content-Disposition'])

        url = reverse('portfolio_word')
        resp = self.client.get(url, {'college': self.co1.id, 'year': 2026})
        self.assertIn('wordprocessingml', resp['Content-Type'])

    def test_quick_exports_scoped_for_college_admin(self):
        from openpyxl import load_workbook

        self.client.login(username="collegeadmin", password="password123")
        resp = self.client.get(reverse('kpi_gap_export'), {'college': self.co2.id})
        self.assertEqual(resp.status_code, 200)
        # Row must reflect CO1's data even though CO2 was requested
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        college_names = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        self.assertIn('Alpha College', college_names)
        self.assertNotIn('Beta College', college_names)

        resp = self.client.get(reverse('submission_status_export'), {'year': 2026})
        self.assertEqual(resp.status_code, 200)
        self.assertIn('submission_status_co1_2026.xlsx', resp['Content-Disposition'])
