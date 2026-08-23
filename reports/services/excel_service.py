"""Excel export of the Annual Portfolio Report (#5)."""
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from su_analytics.constants import MONTH_CHOICES


HEADER_FILL = PatternFill('solid', fgColor='1F3A5F')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color='1F3A5F')
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(bottom=THIN)
NUM_FMT = '#,##0'


def _sheet(wb, title):
    ws = wb.create_sheet(title[:31])
    return ws


def _header_row(ws, row, headers):
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical='center')
    ws.freeze_panes = f'A{row + 1}'


def _autosize(ws):
    for col in range(1, ws.max_column + 1):
        width = 0
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is not None:
                width = max(width, min(len(str(value)) + 2, 60))
        ws.column_dimensions[get_column_letter(col)].width = max(width, 10)


def _num(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value or 0)
    cell.number_format = NUM_FMT
    return cell


def build_portfolio_workbook(context):
    """Render the portfolio context into an xlsx workbook, returned as bytes."""
    chapters = context['chapters']
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Summary ─────────────────────────────────────────────────────
    ws = _sheet(wb, 'Summary')
    ws.cell(row=1, column=1, value=f"{context['college_name']} — Annual Portfolio {chapters['executive_summary']['year']}").font = TITLE_FONT
    _header_row(ws, 3, ['Metric', str(chapters['executive_summary']['prev_year']),
                        str(chapters['executive_summary']['year']), 'Change'])
    for r, m in enumerate(chapters['executive_summary']['yoy'], start=4):
        ws.cell(row=r, column=1, value=m['label'])
        _num(ws, r, 2, m['previous'])
        _num(ws, r, 3, m['current'])
        change = '—' if m['change'] is None else f"{m['change']:+.1f}%"
        ws.cell(row=r, column=4, value=change)
    facts = [
        ('Events held', chapters['executive_summary']['event_count']),
        ('Media coverage items', chapters['executive_summary']['media_count']),
        ('Press releases', chapters['executive_summary']['press_release_count']),
        ('Top posts', chapters['executive_summary']['top_post_count']),
        ('Months reported', chapters['executive_summary']['months_reported']),
        ('KPI targets met', f"{chapters['executive_summary']['kpi_targets_met']} of "
                            f"{chapters['executive_summary']['kpi_target_count']}"),
        ('Avg KPI attainment',
         '—' if chapters['executive_summary']['kpi_avg_attainment'] is None
         else f"{chapters['executive_summary']['kpi_avg_attainment']}%"),
    ]
    start = 4 + len(chapters['executive_summary']['yoy']) + 1
    for i, (label, value) in enumerate(facts):
        ws.cell(row=start + i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=start + i, column=2, value=value)
    _autosize(ws)

    # ── Social media ────────────────────────────────────────────────
    sm = chapters['social_media']
    ws = _sheet(wb, 'Social Media')
    _header_row(ws, 1, ['Month', 'Instagram Views', 'Facebook Views', 'Total Views',
                        'Total Reach', 'Followers Gained', 'Reels', 'Graphics'])
    month_names = dict(MONTH_CHOICES)
    for r, row in enumerate(sm['monthly'], start=2):
        rec = row['record']
        ws.cell(row=r, column=1, value=month_names.get(row['num'], row['num']))
        _num(ws, r, 2, rec.instagram_views)
        _num(ws, r, 3, rec.facebook_views)
        _num(ws, r, 4, rec.total_views)
        _num(ws, r, 5, rec.total_reach)
        _num(ws, r, 6, rec.followers_gained)
        _num(ws, r, 7, rec.reels_count)
        _num(ws, r, 8, rec.graphics_count)
    total_row = 2 + len(sm['monthly'])
    ws.cell(row=total_row, column=1, value='Total').font = Font(bold=True)
    # Totals come from the year aggregate so they match the Summary sheet
    _num(ws, total_row, 4, sm['totals']['total_views'])
    _num(ws, total_row, 5, sm['totals']['total_reach'])
    _num(ws, total_row, 6, sm['totals']['followers_gained'])
    _num(ws, total_row, 7, sm['totals']['reels_count'])
    _num(ws, total_row, 8, sm['totals']['graphics_count'])
    for col in range(2, 9):
        ws.cell(row=total_row, column=col).font = Font(bold=True)

    post_start = total_row + 2
    if sm['top_posts']:
        ws.cell(row=post_start, column=1, value='Top Posts').font = TITLE_FONT
        _header_row(ws, post_start + 1, ['Month', 'Platform', 'Caption', 'Views', 'Likes', 'Shares'])
        for r, p in enumerate(sm['top_posts'], start=post_start + 2):
            ws.cell(row=r, column=1, value=p.month)
            ws.cell(row=r, column=2, value=p.platform.title())
            ws.cell(row=r, column=3, value=(p.caption or '')[:120])
            _num(ws, r, 4, p.views)
            _num(ws, r, 5, p.likes)
            _num(ws, r, 6, p.shares)
    _autosize(ws)

    # ── Events ──────────────────────────────────────────────────────
    ev = chapters['events']
    ws = _sheet(wb, 'Events')
    _header_row(ws, 1, ['Date', 'Title', 'Category', 'Description'])
    for r, e in enumerate(ev['events'], start=2):
        ws.cell(row=r, column=1, value=e.date.strftime('%d %b %Y') if e.date else '')
        ws.cell(row=r, column=2, value=e.title)
        ws.cell(row=r, column=3, value=(e.category or 'other').replace('_', ' ').title())
        ws.cell(row=r, column=4, value=(e.description or '')[:200])
    _autosize(ws)

    # ── Media coverage ──────────────────────────────────────────────
    mc = chapters['media_coverage']
    ws = _sheet(wb, 'Media Coverage')
    _header_row(ws, 1, ['Type', 'Publication / Channel', 'Date', 'Headline / Programme'])
    r = 2
    for n in mc['newspapers']:
        ws.cell(row=r, column=1, value='Newspaper')
        ws.cell(row=r, column=2, value=n.publication)
        ws.cell(row=r, column=3, value=n.date.strftime('%d %b %Y') if n.date else '')
        ws.cell(row=r, column=4, value=n.headline)
        r += 1
    for c_ in mc['channels']:
        ws.cell(row=r, column=1, value='TV / Channel')
        ws.cell(row=r, column=2, value=c_.channel_name)
        ws.cell(row=r, column=3, value=f"{c_.month}/{c_.year}")
        ws.cell(row=r, column=4, value=c_.platform or '')
        r += 1
    _autosize(ws)

    # ── Press releases ──────────────────────────────────────────────
    pr = chapters['press_releases']
    ws = _sheet(wb, 'Press Releases')
    _header_row(ws, 1, ['Date', 'Title', 'Placements', 'Potential Reach'])
    for r, p in enumerate(pr['releases'], start=2):
        ws.cell(row=r, column=1, value=p.date_submitted.strftime('%d %b %Y') if p.date_submitted else '')
        ws.cell(row=r, column=2, value=p.title)
        _num(ws, r, 3, p.placements)
        ws.cell(row=r, column=4, value=str(p.potential_reach or ''))
    totals_row = 2 + len(pr['releases'])
    ws.cell(row=totals_row, column=2, value='Total placements').font = Font(bold=True)
    _num(ws, totals_row, 3, pr['total_placements']).font = Font(bold=True)
    _autosize(ws)

    # ── KPI performance ─────────────────────────────────────────────
    kpi = chapters['kpi_performance']['rows']
    ws = _sheet(wb, 'KPI Performance')
    _header_row(ws, 1, ['Scope', 'Metric', 'Target', 'Actual', 'Gap', 'Achievement %', 'Status'])
    for r, row in enumerate(kpi, start=2):
        ws.cell(row=r, column=1, value=str(row['scope']))
        ws.cell(row=r, column=2, value=row['metric'])
        _num(ws, r, 3, row['target'])
        _num(ws, r, 4, row['actual'])
        _num(ws, r, 5, row['gap'])
        ws.cell(row=r, column=6, value=row['achievement'])
        ws.cell(row=r, column=7, value='On track' if row['on_track'] else 'Behind')
    _autosize(ws)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
